import openpyxl

book = openpyxl.load_workbook("C:\\Users\\TUL\\Desktop\\selenium_framework\\excel_data\\sel_exc.xlsx")

sheet = book.active
items = sheet.cell(row=1,column=2)

print(items.value)

for i in range(1,sheet.max_row + 1):
    for j in range(1,sheet.max_column  +1):
        print(sheet.cell(row=i,column=j).value)